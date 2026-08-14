"""Garment catalog import parser tests."""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from pathlib import Path

import pytest

from app.models.enums import GarmentCategory, GarmentServiceType
from app.services.garment_catalog_import import (
    build_import_summary,
    build_template_xlsx,
    normalize_garment_category,
    parse_import_file,
)

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
DEFAULT_XLS = FIXTURES / "default_garment_catalog.xls"


def test_garment_category_normalization() -> None:
    assert normalize_garment_category("Men") == GarmentCategory.men
    assert normalize_garment_category("HOUSEHOLD") == GarmentCategory.household
    assert normalize_garment_category("Other") == GarmentCategory.others
    assert normalize_garment_category("invalid") is None


def test_parse_default_xls_fixture() -> None:
    if not DEFAULT_XLS.exists():
        pytest.skip("default_garment_catalog.xls fixture missing")
    raw = DEFAULT_XLS.read_bytes()
    result = parse_import_file(raw, "default_garment_catalog.xls")
    assert len(result.valid_rows) == 313
    assert result.error_rows == []
    first = result.valid_rows[0]
    assert first.garment_code == "TF"
    assert first.name == "T Shirt"
    assert first.category == GarmentCategory.men
    assert first.is_visible is True
    assert result.valid_rows[0].rates[GarmentServiceType.dry_cleaning] == Decimal("59.00")
    assert result.valid_rows[0].rates[GarmentServiceType.steam_press] == Decimal("15.00")


def test_parse_default_xls_hidden_row() -> None:
    if not DEFAULT_XLS.exists():
        pytest.skip("default_garment_catalog.xls fixture missing")
    raw = DEFAULT_XLS.read_bytes()
    result = parse_import_file(raw, "default_garment_catalog.xls")
    vest = next(row for row in result.valid_rows if row.garment_code == "VS")
    assert vest.is_visible is False


def test_zero_price_skips_rate() -> None:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["T_ISSHOWITEM", "Category", "Garment", "GarmentCode", "Dry Cleaning", "Steam Press"])
    writer.writerow([1, "Men", "T Shirt", "TF", 0, 15])
    result = parse_import_file(buf.getvalue().encode("utf-8"), "sample.csv")
    assert len(result.valid_rows) == 1
    assert GarmentServiceType.dry_cleaning not in result.valid_rows[0].rates
    assert result.valid_rows[0].rates[GarmentServiceType.steam_press] == Decimal("15.00")


def test_invalid_category_is_error_row() -> None:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Category", "Garment", "GarmentCode", "Dry Cleaning"])
    writer.writerow(["Unknown", "T Shirt", "TF", 59])
    result = parse_import_file(buf.getvalue().encode("utf-8"), "bad.csv")
    assert result.valid_rows == []
    assert len(result.error_rows) == 1
    assert "Unknown category" in result.error_rows[0].errors[0]


def test_duplicate_code_last_row_wins() -> None:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Category", "Garment", "GarmentCode", "Dry Cleaning"])
    writer.writerow(["Men", "Old Shirt", "TF", 50])
    writer.writerow(["Men", "New Shirt", "TF", 60])
    result = parse_import_file(buf.getvalue().encode("utf-8"), "dup.csv")
    assert len(result.valid_rows) == 1
    assert result.valid_rows[0].name == "New Shirt"
    assert result.valid_rows[0].rates[GarmentServiceType.dry_cleaning] == Decimal("60.00")
    assert any("Superseded" in err.errors[0] for err in result.error_rows)


def test_build_import_summary_create_update_counts() -> None:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Category", "Garment", "GarmentCode", "Dry Cleaning"])
    writer.writerow(["Men", "T Shirt", "TF", 59])
    writer.writerow(["Men", "Jeans", "Je", 79])
    parsed = parse_import_file(buf.getvalue().encode("utf-8"), "summary.csv")
    summary = build_import_summary(parsed, existing_codes={"tf"})
    assert summary["valid_count"] == 2
    assert summary["create_count"] == 1
    assert summary["update_count"] == 1


def test_build_template_xlsx_has_headers() -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    data = build_template_xlsx()
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers[0] == "T_ISSHOWITEM"
    assert "Dry Cleaning" in headers
    assert "GarmentCode" in headers
    wb.close()
