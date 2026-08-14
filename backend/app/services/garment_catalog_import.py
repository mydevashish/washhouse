"""Parse partner garment catalog bulk import files (Default.xls format)."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable

from app.models.enums import GarmentCategory, GarmentServiceType

TEMPLATE_HEADERS: tuple[str, ...] = (
    "T_ISSHOWITEM",
    "Category",
    "Garment",
    "GarmentCode",
    "COMMERCIAL SERVICE",
    "Dry Cleaning",
    "EXPRESS SERVICE",
    "On Hanger",
    "LINT REMOVER",
    "PREMIUM LAUNDRY",
    "SHOE CLEANING",
    "Steam Press",
    "Starch",
    "Wash and Fold",
    "Wash N Iron",
)

_SERVICE_HEADER_MAP: dict[str, GarmentServiceType] = {
    "commercial service": GarmentServiceType.commercial_service,
    "dry cleaning": GarmentServiceType.dry_cleaning,
    "express service": GarmentServiceType.express_service,
    "on hanger": GarmentServiceType.on_hanger,
    "lint remover": GarmentServiceType.lint_remover,
    "premium laundry": GarmentServiceType.premium_laundry,
    "shoe cleaning": GarmentServiceType.shoe_cleaning,
    "steam press": GarmentServiceType.steam_press,
    "starch": GarmentServiceType.starch,
    "wash and fold": GarmentServiceType.wash_and_fold,
    "wash n iron": GarmentServiceType.wash_n_iron,
}

_CATEGORY_MAP: dict[str, GarmentCategory] = {
    "men": GarmentCategory.men,
    "women": GarmentCategory.women,
    "kids": GarmentCategory.kids,
    "household": GarmentCategory.household,
    "institutional": GarmentCategory.institutional,
    "others": GarmentCategory.others,
    "other": GarmentCategory.others,
}

_GARMENT_CODE_RE = re.compile(r"^[A-Za-z0-9_-]{1,20}$")


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_visible(raw: Any) -> bool:
    text = _cell_str(raw).lower()
    return text in {"1", "1.0", "true", "yes", "y"}


def _parse_price(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    text = _cell_str(raw)
    if not text or text in {"0", "0.0", "0.00"}:
        return None
    try:
        value = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError(f"Invalid price '{text}'") from None
    if value < 0:
        raise ValueError(f"Price cannot be negative ({text})")
    return value.quantize(Decimal("0.01"))


def normalize_garment_category(raw: str) -> GarmentCategory | None:
    return _CATEGORY_MAP.get(raw.strip().lower())


@dataclass(frozen=True)
class ParsedGarmentRow:
    row_number: int
    garment_code: str
    name: str
    category: GarmentCategory
    is_visible: bool
    rates: dict[GarmentServiceType, Decimal] = field(default_factory=dict)


@dataclass(frozen=True)
class ImportRowError:
    row_number: int
    garment_code: str | None
    name: str | None
    errors: list[str]


@dataclass(frozen=True)
class ImportParseResult:
    valid_rows: list[ParsedGarmentRow]
    error_rows: list[ImportRowError]

    @property
    def total_rows(self) -> int:
        return len(self.valid_rows) + len(self.error_rows)


def _iter_sheet_rows(file_bytes: bytes, filename: str) -> Iterable[list[Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            yield row
        return
    if lower.endswith(".xls"):
        import xlrd

        book = xlrd.open_workbook(file_contents=file_bytes)
        sheet = book.sheet_by_index(0)
        for r in range(sheet.nrows):
            yield [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        return
    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield list(row)
        wb.close()
        return
    raise ValueError("Unsupported file type. Upload .xls, .xlsx, or .csv")


def _build_column_map(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        key = _norm_header(header)
        if not key:
            continue
        mapping[key] = idx
        if key == "garmentcode":
            mapping["garment_code"] = idx
        if key == "t_isshowitem":
            mapping["is_visible"] = idx
    return mapping


def _required_column(columns: dict[str, int], *names: str) -> int | None:
    for name in names:
        idx = columns.get(name)
        if idx is not None:
            return idx
    return None


def parse_import_file(file_bytes: bytes, filename: str) -> ImportParseResult:
    """Parse bulk upload file into valid rows and row-level errors."""
    rows = list(_iter_sheet_rows(file_bytes, filename))
    if not rows:
        return ImportParseResult([], [ImportRowError(1, None, None, ["File is empty"])])

    columns = _build_column_map(rows[0])
    cat_idx = _required_column(columns, "category")
    name_idx = _required_column(columns, "garment", "name")
    code_idx = _required_column(columns, "garment_code", "garmentcode")
    vis_idx = columns.get("is_visible", columns.get("t_isshowitem"))

    missing = [
        label
        for label, idx in (("Category", cat_idx), ("Garment", name_idx), ("GarmentCode", code_idx))
        if idx is None
    ]
    if missing:
        return ImportParseResult(
            [],
            [ImportRowError(1, None, None, [f"Missing required columns: {', '.join(missing)}"])],
        )

    service_columns: list[tuple[int, GarmentServiceType]] = []
    for norm, service_type in _SERVICE_HEADER_MAP.items():
        idx = columns.get(norm)
        if idx is not None:
            service_columns.append((idx, service_type))

    raw_valid: list[ParsedGarmentRow | ImportRowError] = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not any(_cell_str(cell) for cell in row):
            continue

        errors: list[str] = []
        code = _cell_str(row[code_idx]) if code_idx is not None and code_idx < len(row) else ""
        name = _cell_str(row[name_idx]) if name_idx is not None and name_idx < len(row) else ""
        category_raw = _cell_str(row[cat_idx]) if cat_idx is not None and cat_idx < len(row) else ""

        if not code:
            errors.append("GarmentCode is required")
        elif len(code) > 20:
            errors.append("GarmentCode must be at most 20 characters")
        elif not _GARMENT_CODE_RE.match(code):
            errors.append("GarmentCode must be alphanumeric (hyphens/underscores allowed)")

        if not name:
            errors.append("Garment name is required")
        elif len(name) > 120:
            errors.append("Garment name must be at most 120 characters")

        category = normalize_garment_category(category_raw)
        if not category_raw:
            errors.append("Category is required")
        elif category is None:
            errors.append(
                f"Unknown category '{category_raw}'. "
                f"Use one of: Men, Women, Kids, Household, Institutional, Others",
            )

        is_visible = _parse_visible(row[vis_idx]) if vis_idx is not None and vis_idx < len(row) else True

        rates: dict[GarmentServiceType, Decimal] = {}
        for col_idx, service_type in service_columns:
            if col_idx >= len(row):
                continue
            try:
                price = _parse_price(row[col_idx])
            except ValueError as exc:
                errors.append(f"{service_type.value}: {exc}")
                continue
            if price is not None:
                rates[service_type] = price

        if errors:
            raw_valid.append(
                ImportRowError(
                    row_number=row_num,
                    garment_code=code or None,
                    name=name or None,
                    errors=errors,
                ),
            )
            continue

        assert category is not None
        raw_valid.append(
            ParsedGarmentRow(
                row_number=row_num,
                garment_code=code,
                name=name,
                category=category,
                is_visible=is_visible,
                rates=rates,
            ),
        )

    deduped: dict[str, ParsedGarmentRow] = {}
    superseded: list[ImportRowError] = []
    for entry in raw_valid:
        if isinstance(entry, ImportRowError):
            continue
        key = entry.garment_code.lower()
        previous = deduped.get(key)
        if previous is not None:
            superseded.append(
                ImportRowError(
                    row_number=previous.row_number,
                    garment_code=previous.garment_code,
                    name=previous.name,
                    errors=[f"Superseded by row {entry.row_number} (duplicate GarmentCode)"],
                ),
            )
        deduped[key] = entry

    error_rows = [entry for entry in raw_valid if isinstance(entry, ImportRowError)]
    error_rows.extend(superseded)
    valid_rows = sorted(deduped.values(), key=lambda r: r.row_number)
    return ImportParseResult(valid_rows=valid_rows, error_rows=error_rows)


def build_template_xlsx() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(list(TEMPLATE_HEADERS))
    ws.append([1, "Men", "T Shirt", "TF", None, 59, None, None, None, None, None, 15, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_import_summary(
    result: ImportParseResult,
    *,
    existing_codes: set[str],
) -> dict[str, int]:
    create_count = sum(1 for row in result.valid_rows if row.garment_code.lower() not in existing_codes)
    update_count = len(result.valid_rows) - create_count
    return {
        "total_rows": result.total_rows,
        "valid_count": len(result.valid_rows),
        "error_count": len(result.error_rows),
        "create_count": create_count,
        "update_count": update_count,
    }
